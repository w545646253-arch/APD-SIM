#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Right-click direct APD-SIM-3/6/9 inference for one arbitrary GT image."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
import json
import math
import os
from pathlib import Path
import subprocess
import sys
from typing import Any, Sequence


# USER-EDITABLE CONFIGURATION
PROJECT_ROOT = Path(__file__).resolve().parents[2]s[2]
DEFAULT_GT_PATH = Path(
    r"data/sealed_test_gt\microtubules_Cell_054_SIM_gt.tif"
)
DEFAULT_OUTPUT_ROOT = PROJECT_ROOT / "outputs" / "single_gt_apd369_final_assets"
DEFAULT_SEED = 20260812
DEFAULT_PIXEL_SIZE_UM = 0.030588

BUNDLED_NODE = Path(
    r"data/restricted_input"
)
WORKBOOK_BUILDER = PROJECT_ROOT / "tools" / "apd369_single_gt_workbook_runtime" / "build_frc_workbook.mjs"
BUNDLED_NODE_MODULES = Path(
    r"data/restricted_input"
)
MAIN_TEX_PATH = Path(r"docs/manuscript_candidate\APD-SIM.tex")
FINAL_STATUS = "APD369_SINGLE_GT_FRC_SPECTRA_READY"

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from unisim.apd369_single_gt_repaired import (  # noqa: E402
    SingleGTContractError,
    atomic_json,
    build_single_gt_assets,
    run_single_gt_apd369,
    sha256_file,
)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--gt", type=Path, default=DEFAULT_GT_PATH)
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--pixel-size-um", type=float, default=DEFAULT_PIXEL_SIZE_UM)
    parser.add_argument("--crop-x", type=int, default=None)
    parser.add_argument("--crop-y", type=int, default=None)
    parser.add_argument("--dpi", type=int, default=600)
    return parser.parse_args(argv)


def _ensure_workbook_runtime() -> None:
    if not BUNDLED_NODE.is_file() or not BUNDLED_NODE_MODULES.is_dir():
        raise SingleGTContractError("bundled spreadsheet runtime is unavailable")
    if not WORKBOOK_BUILDER.is_file():
        raise SingleGTContractError(f"workbook builder absent: {WORKBOOK_BUILDER}")
    junction = WORKBOOK_BUILDER.parent / "node_modules"
    if not junction.exists():
        completed = subprocess.run(
            ["cmd.exe", "/d", "/c", "mklink", "/J", str(junction), str(BUNDLED_NODE_MODULES)],
            capture_output=True, text=True, check=False,
        )
        if completed.returncode != 0 or not junction.exists():
            raise SingleGTContractError(f"cannot prepare workbook runtime: {completed.stdout} {completed.stderr}")


def _build_workbook(payload: Path, output: Path, preview_dir: Path, log_path: Path) -> dict[str, Any]:
    _ensure_workbook_runtime()
    command = [
        str(BUNDLED_NODE), str(WORKBOOK_BUILDER),
        "--input", str(payload).replace("\\", "/"),
        "--output", str(output).replace("\\", "/"),
        "--preview-dir", str(preview_dir).replace("\\", "/"),
    ]
    completed = subprocess.run(command, cwd=WORKBOOK_BUILDER.parent, capture_output=True, text=True, check=False)
    log_path.parent.mkdir(parents=True, exist_ok=True)
    log_path.write_text((completed.stdout or "") + (completed.stderr or ""), encoding="utf-8")
    if completed.returncode != 0 or not output.is_file():
        raise SingleGTContractError(f"FRC workbook generation failed; see {log_path}")
    return {
        "builder": str(WORKBOOK_BUILDER.resolve()), "builder_sha256": sha256_file(WORKBOOK_BUILDER),
        "node": str(BUNDLED_NODE), "command": command, "exit_code": completed.returncode,
        "log_path": str(log_path.resolve()), "output_sha256": sha256_file(output),
    }


def _validate_workbook(path: Path, payload_path: Path) -> dict[str, Any]:
    """Read-only openpyxl validation; artifact-tool remains the sole author."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SingleGTContractError("openpyxl is required for read-only workbook validation") from exc
    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    required = ["APD-SIM-3", "APD-SIM-6", "APD-SIM-9", "Combined", "Summary", "Contracts"]
    workbook = load_workbook(path, read_only=True, data_only=True)
    if workbook.sheetnames != required:
        raise SingleGTContractError(f"workbook sheet contract failed: {workbook.sheetnames}")
    invalid_strings, formula_errors = [], []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    if value.strip().lower() in {"nan", "inf", "-inf", "infinity", "-infinity"}:
                        invalid_strings.append(f"{sheet.title}!{cell.coordinate}")
                    if value in {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}:
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}")
    method_checks: dict[str, Any] = {}
    for method in required[:3]:
        sheet = workbook[method]
        raw_values = [sheet.cell(row=row, column=2).value for row in range(5, 105)]
        valid_count = sum(isinstance(value, (int, float)) and math.isfinite(float(value)) for value in raw_values)
        expected_cutoff = payload["methods"][method]["cutoff_f_norm"]
        observed_cutoff = [sheet["D5"].value, sheet["D6"].value]
        observed_cutoff_y = [sheet["E5"].value, sheet["E6"].value]
        if expected_cutoff is None:
            cutoff_ok = observed_cutoff == [None, None] and observed_cutoff_y == [None, None]
        else:
            cutoff_ok = all(math.isclose(float(value), float(expected_cutoff), abs_tol=1e-12) for value in observed_cutoff) and observed_cutoff_y == [0, 1]
        threshold_ok = sheet["F5"].value == 0 and sheet["F6"].value == 1 and all(
            math.isclose(float(value), 1.0 / 7.0, abs_tol=1e-12)
            for value in (sheet["G5"].value, sheet["G6"].value)
        )
        if valid_count < 90 or not cutoff_ok or not threshold_ok:
            raise SingleGTContractError(f"workbook method validation failed: {method}")
        method_checks[method] = {"valid_frc_raw_count": valid_count, "cutoff_marker": "PASS", "threshold_marker": "PASS"}
    workbook.close()
    if invalid_strings or formula_errors:
        raise SingleGTContractError(f"workbook contains invalid cells: {invalid_strings} {formula_errors}")
    return {
        "status": "PASS", "path": str(path.resolve()), "sha256": sha256_file(path),
        "sheet_names": required, "formula_error_count": 0, "nonfinite_string_count": 0,
        "openpyxl_reopen": "PASS", "methods": method_checks,
    }


def _write_output_hashes(run_dir: Path, output: Path) -> None:
    rows = []
    for path in sorted(item for item in run_dir.rglob("*") if item.is_file() and item.resolve() != output.resolve()):
        rows.append({"relative_path": path.relative_to(run_dir).as_posix(), "size_bytes": path.stat().st_size, "sha256": sha256_file(path)})
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["relative_path", "size_bytes", "sha256"], delimiter="\t", lineterminator="\n")
        writer.writeheader(); writer.writerows(rows)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if not math.isfinite(args.pixel_size_um) or args.pixel_size_um <= 0:
        raise SingleGTContractError("--pixel-size-um must be positive and finite")
    if args.dpi <= 0:
        raise SingleGTContractError("--dpi must be positive")
    if (args.crop_x is None) != (args.crop_y is None):
        raise SingleGTContractError("--crop-x and --crop-y must be supplied together")
    crop_xy = None if args.crop_x is None else (int(args.crop_x), int(args.crop_y))
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = args.output_root.expanduser().resolve() / f"{args.gt.stem}_seed{int(args.seed)}_{timestamp}"
    checkpoint_paths = [
        PROJECT_ROOT / "checkpoints" / "apd_dmd_geometry_r2" / "dmd3_restart_simple_r1" / "best.pt",
        PROJECT_ROOT / "checkpoints" / "apd_dmd_geometry_r2" / "dmd6" / "best.pt",
        PROJECT_ROOT / "checkpoints" / "apd_dmd_geometry_r4" / "dmd9" / "best.pt",
    ]
    checkpoint_before = {str(path.resolve()): sha256_file(path) for path in checkpoint_paths}
    tex_before = sha256_file(MAIN_TEX_PATH) if MAIN_TEX_PATH.is_file() else None
    print("APD369 direct single-GT repaired inference")
    print(f"custom GT path: {args.gt.expanduser().resolve()}")
    print(f"run directory: {run_dir}")
    inference = run_single_gt_apd369(args.gt, int(args.seed), crop_xy, run_dir)
    assets = build_single_gt_assets(inference, float(args.pixel_size_um), int(args.dpi))
    workbook_build = _build_workbook(
        assets["workbook_payload"], assets["workbook_path"],
        run_dir / "08_receipts" / "workbook_previews", run_dir / "08_receipts" / "workbook_builder.log",
    )
    workbook_validation = _validate_workbook(assets["workbook_path"], assets["workbook_payload"])
    atomic_json(run_dir / "08_receipts" / "excel_validation.json", workbook_validation)
    checkpoint_after = {str(path.resolve()): sha256_file(path) for path in checkpoint_paths}
    checkpoint_modifications = sum(checkpoint_before[path] != checkpoint_after[path] for path in checkpoint_before)
    tex_after = sha256_file(MAIN_TEX_PATH) if MAIN_TEX_PATH.is_file() else None
    main_tex_modifications = int(tex_before != tex_after)
    contract = inference["contract"]
    receipt = {
        "schema_version": 1, "status": FINAL_STATUS, "created_utc": utc_now(),
        "custom_gt_path": str(inference["prepared"].source_path),
        "prepared_gt_path": str((run_dir / "00_input" / "prepared_gt_normalized.tif").resolve()),
        "crop_xywh": list(inference["prepared"].crop_xywh), "run_directory": str(run_dir),
        "repair_pointer": str(contract.repair_pointer),
        "final_repair_status": "DMD9_REPAIRED_APD369_READY",
        "final_repair_status_path": str(contract.final_status_path),
        "formal_test_rerun_count": 0, "formal_30fov_completion_gate_access_count": 0,
        "legacy_batch_exporter_execution_count": 0, "training_execution_count": 0,
        "checkpoint_modification_count": checkpoint_modifications, "main_tex_modification_count": main_tex_modifications,
        "checkpoint_hashes_before": checkpoint_before, "checkpoint_hashes_after": checkpoint_after,
        "main_tex_path": str(MAIN_TEX_PATH), "main_tex_sha256_before": tex_before, "main_tex_sha256_after": tex_after,
        "protocols": {
            plan.method: {
                "checkpoint_path": str(plan.checkpoint_path), "checkpoint_sha256": plan.checkpoint_sha256,
                "protocol_id": plan.protocol_id, "protocol_hash": plan.protocol_hash,
                "raw_order": list(plan.raw_order), "validity_mask": list(plan.validity_mask),
                "raw_stack_sha256": inference["raw_receipts"][plan.method]["raw_stack_sha256"],
                "measurement_seed": inference["raw_receipts"][plan.method]["measurement_seed"],
                "diffusion_seed": inference["reconstruction_receipts"][plan.method]["diffusion_seed"],
                "stage1": inference["reconstruction_receipts"][plan.method]["stage1"],
            } for plan in contract.plans
        },
        "metrics": assets["metrics"], "realspace_contract": assets["realspace_contract"],
        "spectrum_contract": assets["spectrum_contract"],
        "independent_protocol_forward_call_count": inference["independent_protocol_forward_call_count"],
        "common_nine_frame_subsampling_count": inference["common_nine_frame_subsampling_count"],
        "workbook_build": workbook_build, "workbook_validation": workbook_validation,
        "outputs": {
            "harmonized": str((run_dir / "02_harmonized_float").resolve()),
            "realspace_display": str((run_dir / "04_display_rgb").resolve()),
            "spectra": str((run_dir / "05_spectra").resolve()),
            "intensity_colorbar": str((run_dir / "04_display_rgb" / "intensity_colorbar_255_128_0.png").resolve()),
            "spectrum_energy_colorbar": str((run_dir / "05_spectra" / "spectrum_energy_colorbar_1_0p5_0.png").resolve()),
            "frc_workbook": str(assets["workbook_path"].resolve()), "metrics_csv": str(assets["metrics_path"].resolve()),
        },
    }
    if checkpoint_modifications or main_tex_modifications:
        raise SingleGTContractError("protected checkpoint or main TeX changed")
    final_receipt_path = run_dir / "08_receipts" / "final_receipt.json"
    atomic_json(final_receipt_path, receipt)
    atomic_json(
        args.output_root / "CURRENT.json",
        {
            "run_dir": str(run_dir.resolve()),
            "receipt": str(final_receipt_path.resolve()),
            "status": FINAL_STATUS,
        },
    )
    _write_output_hashes(run_dir, run_dir / "08_receipts" / "output_hashes.tsv")
    print(f"custom GT path: {receipt['custom_gt_path']}")
    print(f"prepared GT path: {receipt['prepared_gt_path']}")
    print(f"crop xywh: {receipt['crop_xywh']}")
    for method in ("APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        protocol = receipt["protocols"][method]
        metric = next(row for row in receipt["metrics"] if row["method"] == method)
        print(f"{method} checkpoint: {protocol['checkpoint_path']} / {protocol['checkpoint_sha256']}")
        print(f"{method} protocol/raw order: {protocol['protocol_id']} / {','.join(protocol['raw_order'])}")
        print(f"{method} PSNR/SSIM/FRC/AUC: {metric['psnr_db']:.6f} / {metric['ssim']:.7f} / {metric['gt_frc_period_px']} / {metric['frc_auc']}")
    print("DMD9 tiled inference: tile=320; core=160; shared spatial noise=true")
    print(f"run directory: {run_dir}")
    print(f"harmonized output directory: {run_dir / '02_harmonized_float'}")
    print(f"real-space display directory: {run_dir / '04_display_rgb'}")
    print(f"spectra directory: {run_dir / '05_spectra'}")
    print(f"intensity colorbar path: {receipt['outputs']['intensity_colorbar']}")
    print(f"spectrum energy colorbar path: {receipt['outputs']['spectrum_energy_colorbar']}")
    print(f"FRC workbook path: {receipt['outputs']['frc_workbook']}")
    print(f"metrics CSV path: {receipt['outputs']['metrics_csv']}")
    print(f"final receipt path: {final_receipt_path}")
    print("generate_apd369_from_gt execution count: 0")
    print("formal 30-FOV completion gate access count: 0")
    print("training execution count: 0")
    print("checkpoint modification count: 0")
    print("main TeX modification count: 0")
    print(FINAL_STATUS)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (SingleGTContractError, OSError, ValueError, KeyError, RuntimeError) as error:
        print(f"[APD369 SINGLE-GT ERROR] {error}", file=sys.stderr)
        raise SystemExit(2)
