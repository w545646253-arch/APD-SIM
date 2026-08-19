from __future__ import annotations

import ast
import hashlib
import json
import math
from pathlib import Path

import numpy as np
from PIL import Image
import pytest

from unisim import apd369_single_gt_repaired as single


ROOT = Path(__file__).resolve().parents[1]
LAUNCHER = ROOT / "run_single_gt_apd369_frc_spectra.py"
BACKEND = ROOT / "unisim" / "apd369_single_gt_repaired.py"
OUTPUT_ROOT = ROOT / "outputs" / "single_gt_apd369_final_assets"
EXPECTED_SHEETS = ["APD-SIM-3", "APD-SIM-6", "APD-SIM-9", "Combined", "Summary", "Contracts"]


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


@pytest.fixture(scope="session")
def contract():
    return single.load_apd369_frozen_contract()


@pytest.fixture(scope="session")
def current():
    pointer = json.loads((OUTPUT_ROOT / "CURRENT.json").read_text(encoding="utf-8"))
    run_dir = Path(pointer["run_dir"])
    receipt = json.loads(Path(pointer["receipt"]).read_text(encoding="utf-8"))
    assert pointer["status"] == "APD369_SINGLE_GT_FRC_SPECTRA_READY"
    assert receipt["status"] == pointer["status"]
    return run_dir, receipt


def _plans(contract):
    return {plan.method: plan for plan in contract.plans}


def test_launcher_never_executes_legacy_batch_exporter():
    source = LAUNCHER.read_text(encoding="utf-8")
    tree = ast.parse(source)
    commands = [ast.get_source_segment(source, node) or "" for node in ast.walk(tree) if isinstance(node, ast.Call)]
    assert not any("generate_apd369_from_gt.py" in command for command in commands)


def test_backend_never_references_legacy_batch_exporter():
    assert "generate_apd369_from_gt.py" not in BACKEND.read_text(encoding="utf-8")


def test_launcher_has_no_test30_manifest_dependency():
    assert "test30_manifest.json" not in LAUNCHER.read_text(encoding="utf-8")
    assert "test30_manifest.json" not in BACKEND.read_text(encoding="utf-8")


def test_removed_formal_cli_arguments_are_absent():
    source = LAUNCHER.read_text(encoding="utf-8")
    assert "--formal-root" not in source
    assert "--inference-script" not in source
    assert "--reuse-inference-dir" not in source


def test_apd3_checkpoint_hash(contract):
    plan = _plans(contract)["APD-SIM-3"]
    assert plan.checkpoint_sha256 == "c749e25f625c88dc5f1fbb84ebb1854f760cad21f087c58ed4bf7e734e3919b5"
    assert _sha256(plan.checkpoint_path) == plan.checkpoint_sha256


def test_apd6_checkpoint_hash(contract):
    plan = _plans(contract)["APD-SIM-6"]
    assert plan.checkpoint_sha256 == "10fb16662a8b71b877f2cab81bdc151dcded92f6efd1c4b006306b901a8adff7"
    assert _sha256(plan.checkpoint_path) == plan.checkpoint_sha256


def test_apd9_checkpoint_hash(contract):
    plan = _plans(contract)["APD-SIM-9"]
    assert plan.checkpoint_sha256 == "62831cc9798c9d005fdbf56b343928cc592646b6e70a16f58399b6da0d01b63e"
    assert _sha256(plan.checkpoint_path) == plan.checkpoint_sha256


def test_old_apd9_r3_is_rejected(contract):
    assert single.FORBIDDEN_DMD9_SHA256 == "e4eb12c32041ba99a44ceb479aae431c3892f35f1408269d8d976d55ddb97c47"
    assert all(plan.checkpoint_sha256 != single.FORBIDDEN_DMD9_SHA256 for plan in contract.plans)


@pytest.mark.parametrize(
    ("method", "expected"),
    [
        ("APD-SIM-3", ("X0", "X120", "X240")),
        ("APD-SIM-6", ("H0", "H120", "H240", "V0", "V120", "V240")),
        ("APD-SIM-9", ("X0", "X120", "X240", "Y0", "Y120", "Y240", "Z0", "Z120", "Z240")),
    ],
)
def test_raw_orders(contract, method, expected):
    assert _plans(contract)[method].raw_order == expected


def test_validity_masks(contract):
    plans = _plans(contract)
    assert plans["APD-SIM-3"].validity_mask == (1, 1, 1) + (0,) * 12
    assert plans["APD-SIM-6"].validity_mask == (1,) * 6 + (0,) * 9
    assert plans["APD-SIM-9"].validity_mask == (1,) * 9 + (0,) * 6


def test_three_raw_stacks_are_independently_generated(current):
    _, receipt = current
    assert receipt["independent_protocol_forward_call_count"] == 3
    assert receipt["common_nine_frame_subsampling_count"] == 0
    raw_receipts = []
    run_dir = Path(receipt["run_directory"])
    for method in ("APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        protocol = receipt["protocols"][method]["protocol_id"]
        raw_receipts.append(json.loads((run_dir / "01_raw_measurements" / protocol / "raw_stack_receipt.json").read_text(encoding="utf-8")))
    assert len({row["generation_call_uuid"] for row in raw_receipts}) == 3
    assert all(row["generation_call_kind"] == "independent_protocol_forward" for row in raw_receipts)
    assert all(row["source_nine_frame_subsampling"] is False for row in raw_receipts)


def test_raw_stack_shapes_match_protocols(current):
    run_dir, receipt = current
    expected = {"APD-SIM-3": 3, "APD-SIM-6": 6, "APD-SIM-9": 9}
    for method, frame_count in expected.items():
        protocol = receipt["protocols"][method]["protocol_id"]
        raw = np.load(run_dir / "01_raw_measurements" / protocol / "raw_stack.npy", mmap_mode="r")
        assert raw.shape == (frame_count, 1004, 1004)
        assert raw.dtype == np.float32
        assert np.isfinite(raw).all()


def test_dmd9_tiled_stage1_was_used(current):
    _, receipt = current
    stage1 = receipt["protocols"]["APD-SIM-9"]["stage1"]
    assert stage1["mode"] == "tiled"
    assert stage1["tile_size"] == 320
    assert stage1["core_size"] == 160
    assert stage1["shared_single_spatial_noise_field"] is True
    assert stage1["deterministic_stitching"] is True


def test_dmd9_single_block_full_size_was_not_used(current):
    _, receipt = current
    assert receipt["protocols"]["APD-SIM-9"]["stage1"]["single_block_1004_inference"] is False


def test_harmonized_output_contract(current):
    run_dir, _ = current
    for name in ("GT", "APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        value = np.load(run_dir / "02_harmonized_float" / f"{name}.npy", mmap_mode="r")
        assert value.shape == (1004, 1004)
        assert value.dtype == np.float32
        assert np.isfinite(value).all()
        assert float(value.min()) >= 0.0 and float(value.max()) <= 1.0


def test_frc_raw_comes_from_formal_gt_frc(current):
    from tools import revision_dmd6_common as common

    run_dir, receipt = current
    arrays = {name: np.load(run_dir / "02_harmonized_float" / f"{name}.npy") for name in ("GT", "APD-SIM-3", "APD-SIM-6", "APD-SIM-9")}
    curves = np.load(run_dir / "06_frc" / "frc_curves.npz")
    for method in ("APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        meta, curve = common.gt_frc(arrays["GT"], arrays[method], pixel_size_um=0.030588)
        key = method.replace("APD-SIM-", "") + "F_FRC_raw"
        np.testing.assert_allclose(curves[key], curve["frc"], rtol=0, atol=0, equal_nan=True)
        recorded = next(row for row in receipt["metrics"] if row["method"] == method)
        assert recorded["frc_auc"] == meta["frc_auc_to_cutoff_or_nyquist"]


def test_frc_smooth_does_not_control_cutoff_or_auc(current):
    run_dir, receipt = current
    payload = json.loads((run_dir / "06_frc" / "workbook_payload.json").read_text(encoding="utf-8"))
    for row in receipt["metrics"]:
        method = row["method"]
        assert payload["methods"][method]["cutoff_f_norm"] == row["cutoff_f_norm_nyq1"]
    assert "fixed 7-bin plotting only; excluded from cutoff and AUC" in dict(payload["contracts"])["FRC smooth restriction"]


def test_excel_nonfinite_values_are_blank(current):
    from openpyxl import load_workbook

    run_dir, _ = current
    workbook = load_workbook(run_dir / "06_frc" / "APD369_FRC_curves.xlsx", read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.strip().lower() in {"nan", "inf", "-inf", "infinity", "-infinity"})
    workbook.close()


def test_excel_sheet_and_origin_header_contract(current):
    from openpyxl import load_workbook

    run_dir, _ = current
    workbook = load_workbook(run_dir / "06_frc" / "APD369_FRC_curves.xlsx", read_only=True, data_only=True)
    assert workbook.sheetnames == EXPECTED_SHEETS
    expected = ["f_norm(Nyq=1)", "FRC_raw", "FRC_smooth", "cutoff_x", "cutoff_y", "threshold_x", "threshold_y", "Ni", "thr_halfbit"]
    for method in EXPECTED_SHEETS[:3]:
        assert [workbook[method].cell(1, column).value for column in range(1, 10)] == expected
        assert [workbook[method].cell(4, column).value for column in range(1, 10)] == ["X1", "Y1", "Y2", "X2", "Y3", "X3", "Y4", "Y5", "Y6"]
    workbook.close()


def test_spectra_use_one_shared_scale(current):
    run_dir, receipt = current
    contract = receipt["spectrum_contract"]
    assert contract["shared"] is True and contract["lut"] == "magma"
    low, high = contract["global_low"], contract["global_high"]
    assert math.isfinite(low) and math.isfinite(high) and high > low
    for name in ("GT", "APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        value = np.load(run_dir / "05_spectra" / "normalized_float" / f"{name}.npy", mmap_mode="r")
        assert value.shape == (1004, 1004) and value.dtype == np.float32
        assert np.isfinite(value).all() and float(value.min()) >= 0 and float(value.max()) <= 1


def test_realspace_uses_one_shared_scale(current):
    run_dir, receipt = current
    assert receipt["realspace_contract"] == {"range": [0.0, 1.0], "lut": "black-blue-cyan-white", "shared": True}
    for name in ("GT", "APD-SIM-3", "APD-SIM-6", "APD-SIM-9"):
        display = np.asarray(Image.open(run_dir / "04_display_rgb" / f"{name}.png"))
        assert display.shape == (1004, 1004, 3) and display.dtype == np.uint8


def test_colorbar_files_open(current):
    run_dir, _ = current
    for path in (
        run_dir / "04_display_rgb" / "intensity_colorbar_255_128_0.png",
        run_dir / "04_display_rgb" / "intensity_colorbar_255_128_0.tif",
        run_dir / "05_spectra" / "spectrum_energy_colorbar_1_0p5_0.png",
        run_dir / "05_spectra" / "spectrum_energy_colorbar_1_0p5_0.tif",
    ):
        with Image.open(path) as image:
            image.verify()


def test_checkpoint_hashes_unchanged_by_run(current):
    _, receipt = current
    assert receipt["checkpoint_modification_count"] == 0
    assert receipt["checkpoint_hashes_before"] == receipt["checkpoint_hashes_after"]


def test_main_tex_hash_unchanged_by_run(current):
    _, receipt = current
    assert receipt["main_tex_modification_count"] == 0
    assert receipt["main_tex_sha256_before"] == receipt["main_tex_sha256_after"]


def test_training_and_formal_rerun_counts_are_zero(current):
    _, receipt = current
    assert receipt["training_execution_count"] == 0
    assert receipt["formal_test_rerun_count"] == 0
    assert receipt["formal_30fov_completion_gate_access_count"] == 0
    assert receipt["legacy_batch_exporter_execution_count"] == 0


def test_seed_policy_has_no_protocol_offsets(current):
    _, receipt = current
    measurement = {receipt["protocols"][method]["measurement_seed"] for method in ("APD-SIM-3", "APD-SIM-6", "APD-SIM-9")}
    diffusion = {receipt["protocols"][method]["diffusion_seed"] for method in ("APD-SIM-3", "APD-SIM-6", "APD-SIM-9")}
    assert measurement == {20260812}
    assert diffusion == {20260812}


def test_input_receipt_records_no_resize_or_interpolation(current):
    run_dir, _ = current
    preparation = json.loads((run_dir / "00_input" / "input_preparation_receipt.json").read_text(encoding="utf-8"))
    assert preparation["resize"] is False
    assert preparation["interpolation"] is False
    assert preparation["crop_coordinates_xywh"] == [0, 0, 1004, 1004]


def test_required_csv_npz_and_receipts_exist(current):
    run_dir, _ = current
    required = [
        run_dir / "06_frc" / "frc_curves_long.csv",
        run_dir / "06_frc" / "frc_curves_wide.csv",
        run_dir / "06_frc" / "frc_curves.npz",
        run_dir / "07_metrics" / "single_image_metrics.csv",
        run_dir / "08_receipts" / "final_receipt.json",
        run_dir / "08_receipts" / "excel_validation.json",
    ]
    assert all(path.is_file() and path.stat().st_size > 0 for path in required)
